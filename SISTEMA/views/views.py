from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, DetailView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth import views as auth_views
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.template.loader import get_template
from django.contrib.auth.decorators import login_required


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from collections import Counter


from SISTEMA.forms.forms import LoginForm, PosibleClienteForm, InformeVendedorForm, ClienteForm, ReporteExportForm
from SISTEMA.models.models import InformeVendedor, PosibleCliente, Cliente, Usuario



class SoloAdminMixin(UserPassesTestMixin):
    def test_func(self):
        usuario = self.request.user
        
        if not usuario.is_authenticated:
            return False

        es_admin_tecnico = usuario.is_staff or usuario.is_superuser
        
        es_jefe = usuario.rol is not None and usuario.rol.nombre == "Jefe"

        return es_admin_tecnico or es_jefe

    def handle_no_permission(self):
        messages.error(self.request, "No tienes permiso para acceder a esta opción.")
        return redirect('home')



class LoginView(auth_views.LoginView):
    template_name = 'login.html'
    authentication_form = LoginForm
    redirect_authenticated_user = True
    
    
    
class LogoutView(auth_views.LogoutView):
    next_page = reverse_lazy('login')



class HomeView(LoginRequiredMixin, TemplateView):
    template_name = 'home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cant_reportes'] = InformeVendedor.objects.count()
        context['cant_posibles_clientes'] = PosibleCliente.objects.count()
        return context
    
    
    
class IngresarReporteView(LoginRequiredMixin, CreateView):
    model = InformeVendedor
    form_class = InformeVendedorForm
    template_name = 'IngresarReporte.html'
    success_url = reverse_lazy('ver_reporte')
    
    def form_valid(self, form):
        form.instance.vendedor = self.request.user
        return super().form_valid(form)
    
    

class VerReporteView(LoginRequiredMixin, ListView):
    model = InformeVendedor
    template_name = 'VerReporte.html'
    ordering = ['-fecha_ingreso']
    


class IngresarPosibleClienteView(LoginRequiredMixin, CreateView):
    model = PosibleCliente
    form_class = PosibleClienteForm
    template_name = 'IngresarPosibleCliente.html'
    success_url = reverse_lazy('ver_posible_cliente')
    
    
class VerPosibleClienteView(LoginRequiredMixin, ListView):
    model = PosibleCliente
    template_name = 'VerPosibleCliente.html'
    ordering = ['nombre']
    
    
    
class ConvertirPosibleClienteView(LoginRequiredMixin, CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'IngresarCliente.html'
    success_url = reverse_lazy('ver_cliente')

    def dispatch(self, request, *args, **kwargs):
        self.posible_cliente = get_object_or_404(PosibleCliente, pk=self.kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        initial['rut'] = self.posible_cliente.rut
        initial['nombre_comercio'] = self.posible_cliente.nombre_comercio
        initial['ubicacion'] = self.posible_cliente.ubicacion
        initial['nombre'] = self.posible_cliente.nombre
        initial['apellido'] = self.posible_cliente.apellido
        initial['email'] = self.posible_cliente.email
        initial['telefono'] = self.posible_cliente.telefono
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = f"Convertir Prospecto: {self.posible_cliente.nombre}"
        return context

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            self.posible_cliente.delete()
            
        return redirect(self.success_url)
    
    

class VerClienteView(LoginRequiredMixin, SoloAdminMixin, ListView):
    model = Cliente
    template_name = 'VerCliente.html'
    ordering = ['nombre']
    


class ExportarReportesView(LoginRequiredMixin, FormView):
    template_name = 'ExportarReporte.html'
    form_class = ReporteExportForm

    def form_valid(self, form):
        # 1. Obtener rango de fechas del formulario
        f_inicio = form.cleaned_data['fecha_inicio']
        f_termino = form.cleaned_data['fecha_termino']

        # 2. Preparar los datos
        data_reporte = []

        # PASO CRÍTICO: Buscar IDs de CUALQUIER usuario que tenga reportes en este rango.
        # Esto soluciona que no aparezcan tus reportes de Admin.
        vendedores_ids = InformeVendedor.objects.filter(
            fecha_ingreso__date__range=[f_inicio, f_termino]
        ).values_list('vendedor', flat=True).distinct()
        
        # Traemos los objetos Usuario correspondientes
        vendedores = Usuario.objects.filter(id__in=vendedores_ids)

        for vendedor in vendedores:
            # Buscamos los reportes de ese vendedor específico
            # OPTIMIZACIÓN: Usamos prefetch_related para los productos (Many-to-Many)
            reportes = InformeVendedor.objects.filter(
                vendedor=vendedor,
                fecha_ingreso__date__range=[f_inicio, f_termino]
            ).select_related(
                'cliente', 
                'metodo_pago'
            ).prefetch_related(
                'tipo_producto' # <--- Clave para que los productos carguen rápido en el PDF
            ).order_by('fecha_ingreso')

            if reportes.exists():
                data_reporte.append({
                    'vendedor': vendedor,
                    'reportes': reportes
                })

        # 3. Detectar qué botón presionó el usuario
        if 'btn_excel' in self.request.POST:
            return self.generar_excel(data_reporte, f_inicio, f_termino)
        elif 'btn_pdf' in self.request.POST:
            return self.generar_pdf(data_reporte, f_inicio, f_termino)
        
        return super().form_valid(form)

    # --- GENERADOR DE EXCEL ---
    def generar_excel(self, data_reporte, f_inicio, f_termino):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=Reporte_Ventas_{f_inicio}_{f_termino}.xlsx'

        wb = openpyxl.Workbook()
        
        # ---------------------------------------------------------
        # HOJA 1: DETALLE (La que ya tenías)
        # ---------------------------------------------------------
        ws = wb.active
        ws.title = "Detalle Gestión"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        vendedor_font = Font(bold=True, size=12)
        vendedor_fill = PatternFill(start_color="D9D9D9", fill_type="solid")

        # Variables para acumular datos estadísticos mientras escribimos el detalle
        stats_total_reportes = 0
        stats_total_ventas = 0
        stats_productos = Counter() # Contador inteligente para los productos

        row_num = 1
        for item in data_reporte:
            vendedor = item['vendedor']
            
            # Título Vendedor
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
            cell = ws.cell(row=row_num, column=1, value=f"VENDEDOR: {vendedor.username.upper()}")
            cell.font = vendedor_font
            cell.fill = vendedor_fill
            cell.alignment = Alignment(horizontal='left', vertical='center') 
            row_num += 1

            # Encabezados
            headers = ['Fecha', 'Cliente', 'Comercio', 'Productos', 'Venta?', 'Pago', 'Nota']
            for col, title in enumerate(headers, 1):
                c = ws.cell(row=row_num, column=col, value=title)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal='center')
            row_num += 1

            # Filas
            for rep in item['reportes']:
                # --- RECOLECCIÓN DE ESTADÍSTICAS ---
                stats_total_reportes += 1
                if rep.venta_realizada:
                    stats_total_ventas += 1
                    # Contamos los productos SOLO si se concretó la venta
                    for prod in rep.tipo_producto.all():
                        stats_productos[prod.nombre] += 1
                # -----------------------------------

                # Lógica de validación (Cliente/Pago)
                if rep.cliente:
                    nombre_cli = f"{rep.cliente.nombre} {rep.cliente.apellido}"
                    comercio_cli = rep.cliente.nombre_comercio or "Particular"
                else:
                    nombre_cli = f"{rep.nombre or ''} {rep.apellido or ''}".strip() or "Anónimo"
                    comercio_cli = rep.nombre_comercio or "Particular"

                pago_str = "-"
                if rep.venta_realizada:
                    pago_str = rep.metodo_pago.nombre if rep.metodo_pago else "No especificado"

                productos_str = ", ".join([p.nombre for p in rep.tipo_producto.all()]) or "-"

                row = [
                    rep.fecha_ingreso.strftime('%d/%m/%Y %H:%M'),
                    nombre_cli,
                    comercio_cli,
                    productos_str,
                    "SÍ" if rep.venta_realizada else "NO",
                    pago_str,
                    rep.descripcion
                ]
                
                for col, val in enumerate(row, 1):
                    ws.cell(row=row_num, column=col, value=val)
                row_num += 1
            
            row_num += 2 

        # Ajuste columnas Hoja 1
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50: adjusted_width = 50 
            ws.column_dimensions[column_letter].width = adjusted_width

        # ---------------------------------------------------------
        # HOJA 2: ESTADÍSTICAS (NUEVO)
        # ---------------------------------------------------------
        ws_stats = wb.create_sheet("Estadísticas")
        
        # 1. TABLA DE EFECTIVIDAD DE VENTAS
        ws_stats['A1'] = "RESUMEN GLOBAL"
        ws_stats['A1'].font = Font(bold=True, size=14)
        
        ws_stats['A3'] = "Total Visitas/Reportes"
        ws_stats['B3'] = stats_total_reportes
        
        ws_stats['A4'] = "Ventas Concretadas"
        ws_stats['B4'] = stats_total_ventas
        
        ws_stats['A5'] = "% Efectividad de Venta"
        # Cálculo para evitar división por cero
        tasa_efectividad = (stats_total_ventas / stats_total_reportes) if stats_total_reportes > 0 else 0
        cell_eff = ws_stats['B5']
        cell_eff.value = tasa_efectividad
        cell_eff.number_format = '0.00%' # Formato porcentaje Excel
        
        # Estilo tabla resumen
        for row in range(3, 6):
            ws_stats[f'A{row}'].font = Font(bold=True)
            ws_stats[f'A{row}'].fill = PatternFill(start_color="E2EFDA", fill_type="solid") # Verde claro

        # 2. TABLA DE PRODUCTOS VENDIDOS
        ws_stats['D1'] = "DETALLE POR PRODUCTO"
        ws_stats['D1'].font = Font(bold=True, size=14)

        # Encabezados Tabla Productos
        headers_prod = ['Producto', 'Cant. Vendida', '% del Total Productos']
        cols_prod = ['D', 'E', 'F']
        
        for col_letter, title in zip(cols_prod, headers_prod):
            cell = ws_stats[f'{col_letter}3']
            cell.value = title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Datos Productos
        total_items_vendidos = sum(stats_productos.values())
        row_stats = 4
        
        # Ordenamos los productos del más vendido al menos vendido
        for producto, cantidad in stats_productos.most_common():
            # Nombre
            ws_stats[f'D{row_stats}'] = producto
            
            # Cantidad
            ws_stats[f'E{row_stats}'] = cantidad
            ws_stats[f'E{row_stats}'].alignment = Alignment(horizontal='center')
            
            # Porcentaje
            porcentaje_prod = (cantidad / total_items_vendidos) if total_items_vendidos > 0 else 0
            cell_perc = ws_stats[f'F{row_stats}']
            cell_perc.value = porcentaje_prod
            cell_perc.number_format = '0.00%'
            
            row_stats += 1

        # Ajuste ancho columnas estadísticas
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['D'].width = 30
        ws_stats.column_dimensions['E'].width = 15
        ws_stats.column_dimensions['F'].width = 20

        wb.save(response)
        return response

    # --- GENERADOR DE PDF ---
    def generar_pdf(self, data_reporte, f_inicio, f_termino):
        template_path = 'PdfTemplate.html' # Asegúrate que coincida mayúsculas/minúsculas con tu archivo
        context = {
            'data_reporte': data_reporte,
            'f_inicio': f_inicio,
            'f_termino': f_termino,
            'user': self.request.user,
        }
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_{f_inicio}.pdf"'
        
        template = get_template(template_path)
        html = template.render(context)

        pisa_status = pisa.CreatePDF(html, dest=response)

        if pisa_status.err:
            return HttpResponse('Error al generar PDF: <pre>' + html + '</pre>')
        return response
    
    
def error_404_view(request, exception):
    return render(request, 'Error.html', {
        'mensaje': 'La página que buscas no existe o ha sido movida.'
    }, status=404)

def error_500_view(request):
    return render(request, 'Error.html', {
        'mensaje': 'Tuvimos un problema interno. El equipo técnico ha sido notificado.'
    }, status=500)
    
    
@login_required
def home(request):
    # 1. Obtener los últimos 3 reportes del usuario actual
    # Ordenamos por fecha_inicio descendente ('-') y tomamos los primeros 3 ([:3])
    ultimos_reportes = InformeVendedor.objects.filter(usuario=request.user).order_by('-fecha_inicio')[:3]
    
    context = {
        'ultimos_reportes': ultimos_reportes
    }
    
    return render(request, 'SISTEMA/home.html', context)